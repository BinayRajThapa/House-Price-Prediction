import pandas as pd
from openpyxl import load_workbook

wb_p='Data/DataSet.xlsx'

df=pd.read_excel(wb_p)

#REMOVE NULL 
df.drop(columns=["PARKING","BUILDUP AREA"],inplace=True)
df.dropna(axis=0,inplace=True)

def conv2string(name):
    df[name]=df[name].astype(str)


def city_seperate(old,new):
    seperated=row[old].split(',')
    df.at[index,new]=seperated[-1]

def by_seperate(old,new):
    seperated=row[old].split()
    df.at[index,new]=seperated[0]

def land_seperate(old,new):
    
    seperated=row[old].split()
    if seperated[-1]=="katha":
        df.at[index,new]=float(seperated[0])*10.65
    else:    
        df.at[index,new]=(seperated[0])     

def road_seperate(old,new):
    seperated=row[old].split()
    if seperated[-1]=="Meter":
        df.at[index,new]=float(seperated[0])*3.28084
    else:    
        df.at[index,new]=seperated[0]

def price_seperate(old,new):
    seperate_array=row[old].split()
    if seperate_array[0]=="Rs.":
        numeric_part=seperate_array[1].replace(',','')
        if seperate_array[-1]=="Cr":
            df.at[index,new]=float(numeric_part)*10000000
        else:
            df.at[index,new]=float(numeric_part)
    else:
        df.drop(index,inplace=True)
   

conv2string("LAND AREA")
conv2string("ROAD ACCESS")
conv2string("BUILT YEAR")
conv2string("PRICE")

for index,row in df.iterrows():
    city_seperate("LOCATION","CITY")
    land_seperate("LAND AREA","LA_N")
    road_seperate("ROAD ACCESS","RA_N")
    by_seperate("BUILT YEAR","BY_N")
    price_seperate("PRICE","PRICE_N")
    
workbook=load_workbook(wb_p)
workbook.create_sheet('FilteredData')

sheet=workbook['FilteredData']


df.to_csv("data.csv")


